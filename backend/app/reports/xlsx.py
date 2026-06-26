"""Generate XLSX (Excel) export for transaction data."""

from datetime import datetime
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.validators import format_datetime

HEADER_FILL = PatternFill(start_color="F26822", end_color="F26822", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
META_FONT = Font(size=10, color="666666")


def generate_xlsx_export(
    transactions: List[dict],
    fund_label: str,
    datetime_format: str | None = None,
) -> bytes:
    """Generate an XLSX file from filtered transaction data.

    Columns: txid, amount_atomic, amount_xmr, confirmations, timestamp, unlock_time, height.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Title row
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Transaction Export — {fund_label}"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:G2")
    ws["A2"] = (
        format_datetime(datetime.now(), datetime_format)
        if datetime_format
        else datetime.now().isoformat()
    )
    ws["A2"].font = META_FONT

    # Headers
    headers = [
        "txid",
        "amount_atomic",
        "amount_xmr",
        "confirmations",
        "timestamp",
        "unlock_time",
        "height",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, tx in enumerate(transactions, 5):
        ws.cell(row=row_idx, column=1, value=tx["txid"])
        ws.cell(row=row_idx, column=2, value=tx["amount_atomic"])
        ws.cell(row=row_idx, column=3, value=str(tx["amount_xmr"]))
        ws.cell(row=row_idx, column=4, value=tx["confirmations"])
        ts = tx["timestamp"]
        if datetime_format and hasattr(ts, "year"):
            ws.cell(row=row_idx, column=5, value=format_datetime(ts, datetime_format))
        elif hasattr(ts, "isoformat"):
            ws.cell(row=row_idx, column=5, value=ts.isoformat())
        else:
            ws.cell(row=row_idx, column=5, value=str(ts))
        ws.cell(row=row_idx, column=6, value=tx.get("unlock_time") or 0)
        ws.cell(row=row_idx, column=7, value=tx["height"])

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = len(headers[col_idx - 1])
        for row in ws.iter_rows(
            min_row=5, max_row=5 + len(transactions), min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
